import os
import sqlite3
import openpyxl
from openpyxl import load_workbook
import re

# Создаём базу данных
conn = sqlite3.connect('royalties.db')
cursor = conn.cursor()

# Создаём таблицу
cursor.execute('''
CREATE TABLE IF NOT EXISTS royalties (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract TEXT,
    quarter TEXT,
    year INTEGER,
    type TEXT,
    song TEXT,
    authors TEXT,
    composers TEXT,
    artist TEXT,
    sum REAL,
    display_name TEXT,
    additional_info TEXT
)
''')
conn.commit()

# Функция для парсинга названия файла
def parse_filename(filename):
    # Пример: 955-р 4кв 2025.xlsx
    name = filename.replace('.xlsx', '')
    parts = name.split(' ')
    
    # Извлекаем квартал (1кв, 2кв, 3кв, 4кв или I, II, III, IV)
    quarter_raw = ''
    year_raw = ''
    contract = ''
    
    for part in parts:
        if 'кв' in part:
            quarter_raw = part
        elif part.isdigit() and len(part) == 4:
            year_raw = part
        else:
            contract = part
    
    # Преобразуем квартал в формат I, II, III, IV
    quarter_map = {
        '1кв': 'I', '2кв': 'II', '3кв': 'III', '4кв': 'IV',
        'I': 'I', 'II': 'II', 'III': 'III', 'IV': 'IV'
    }
    quarter = quarter_map.get(quarter_raw, quarter_raw)
    
    # Год
    year = int(year_raw) if year_raw else 0
    
    return contract, quarter, year

# Функция для обработки листа "А" (Авторские)
def process_sheet_a(ws, contract, quarter, year):
    data = []
    
    # Проходим по строкам, начиная с 3-й (пропускаем заголовки)
    for row in ws.iter_rows(min_row=3, max_col=18, values_only=True):
        if not row or not row[0]:
            continue
        
        song = str(row[0]).strip() if row[0] else ''
        authors = str(row[1]).strip() if row[1] else ''
        composers = str(row[2]).strip() if row[2] else ''
        # Сумма в колонке K (индекс 10)
        amount = row[10] if row[10] and isinstance(row[10], (int, float)) else 0
        
        if song and amount > 0:
            data.append({
                'contract': contract,
                'quarter': quarter,
                'year': year,
                'type': 'Авторские',
                'song': song,
                'authors': authors,
                'composers': composers,
                'artist': '',
                'sum': amount,
                'display_name': song,
                'additional_info': f'Авторы: {authors}' if authors else (f'Композиторы: {composers}' if composers else '')
            })
    
    return data

# Функция для обработки листа "С" (Смежные)
def process_sheet_c(ws, contract, quarter, year):
    data = []
    
    # Проходим по строкам, начиная с 3-й
    for row in ws.iter_rows(min_row=3, max_col=18, values_only=True):
        if not row or not row[1]:
            continue
        
        song = str(row[1]).strip() if row[1] else ''
        artist = str(row[2]).strip() if row[2] else ''
        # Сумма в колонке K (индекс 10)
        amount = row[10] if row[10] and isinstance(row[10], (int, float)) else 0
        
        if song and amount > 0:
            data.append({
                'contract': contract,
                'quarter': quarter,
                'year': year,
                'type': 'Смежные',
                'song': song,
                'authors': '',
                'composers': '',
                'artist': artist,
                'sum': amount,
                'display_name': song,
                'additional_info': f'Исполнитель: {artist}' if artist else ''
            })
    
    return data

# Основной процесс загрузки
print("Начинаю загрузку данных...")

# Папка с Excel-файлами
data_folder = 'data'
total_rows = 0

for filename in os.listdir(data_folder):
    if not filename.endswith('.xlsx'):
        continue
    
    print(f"Обрабатываю: {filename}")
    
    # Парсим имя файла
    contract, quarter, year = parse_filename(filename)
    print(f"  Договор: {contract}, Квартал: {quarter}, Год: {year}")
    
    # Открываем файл
    filepath = os.path.join(data_folder, filename)
    wb = load_workbook(filepath, data_only=True)
    
    # Обрабатываем лист "А"
    if 'А' in wb.sheetnames:
        ws = wb['А']
        author_data = process_sheet_a(ws, contract, quarter, year)
        for row in author_data:
            cursor.execute('''
                INSERT INTO royalties (contract, quarter, year, type, song, authors, composers, artist, sum, display_name, additional_info)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (row['contract'], row['quarter'], row['year'], row['type'], 
                  row['song'], row['authors'], row['composers'], row['artist'],
                  row['sum'], row['display_name'], row['additional_info']))
        total_rows += len(author_data)
        print(f"  Добавлено авторских: {len(author_data)}")
    
    # Обрабатываем лист "С"
    if 'С' in wb.sheetnames:
        ws = wb['С']
        related_data = process_sheet_c(ws, contract, quarter, year)
        for row in related_data:
            cursor.execute('''
                INSERT INTO royalties (contract, quarter, year, type, song, authors, composers, artist, sum, display_name, additional_info)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (row['contract'], row['quarter'], row['year'], row['type'], 
                  row['song'], row['authors'], row['composers'], row['artist'],
                  row['sum'], row['display_name'], row['additional_info']))
        total_rows += len(related_data)
        print(f"  Добавлено смежных: {len(related_data)}")
    
    wb.close()

# Сохраняем изменения
conn.commit()
conn.close()

print(f"\n✅ Готово! Загружено {total_rows} записей в базу данных royalties.db")
