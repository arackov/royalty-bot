import os
import sqlite3
import asyncio
import threading
from flask import Flask
from aiogram import Bot, Dispatcher, types
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

load_dotenv()

API_TOKEN = os.getenv("TG_BOT_API_KEY")
ADMIN_ID = int(os.getenv("ADMIN_ID"))
ALLOWED_USERS = [int(x.strip()) for x in os.getenv("ALLOWED_USERS", "").split(",") if x.strip()]

bot = Bot(token=API_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Flask для health check
flask_app = Flask(__name__)

@flask_app.route('/')
@flask_app.route('/health')
def health():
    return "Bot is running", 200

# Состояния для FSM
class ReportState(StatesGroup):
    waiting_contracts = State()
    waiting_quarters = State()
    waiting_years = State()
    waiting_types = State()
    waiting_songs = State()
    waiting_author_percent = State()
    waiting_related_percent = State()

# Вспомогательная функция для получения данных из БД
def get_db_connection():
    conn = sqlite3.connect('royalties.db')
    conn.row_factory = sqlite3.Row
    return conn

# Получение уникальных значений для фильтров
def get_unique_values(column):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"SELECT DISTINCT {column} FROM royalties WHERE {column} IS NOT NULL AND {column} != '' ORDER BY {column}")
    values = [row[0] for row in cursor.fetchall()]
    conn.close()
    return values

# Получение списка песен
def get_songs():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT display_name FROM royalties ORDER BY display_name")
    songs = [row[0] for row in cursor.fetchall()]
    conn.close()
    return songs

# Клавиатура для множественного выбора
def build_multi_select_keyboard(items, selected_items, prefix, page=0, items_per_page=10):
    keyboard = []
    start_idx = page * items_per_page
    end_idx = min(start_idx + items_per_page, len(items))
    
    for item in items[start_idx:end_idx]:
        is_selected = item in selected_items
        emoji = "✅ " if is_selected else "⬜ "
        callback_data = f"{prefix}_toggle_{item.replace(' ', '_')}"
        keyboard.append([InlineKeyboardButton(text=f"{emoji}{item}", callback_data=callback_data)])
    
    # Навигация
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(text="◀️ Назад", callback_data=f"{prefix}_page_{page-1}"))
    if end_idx < len(items):
        nav_buttons.append(InlineKeyboardButton(text="Вперед ▶️", callback_data=f"{prefix}_page_{page+1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Кнопка "Готово"
    keyboard.append([InlineKeyboardButton(text="✅ Готово", callback_data=f"{prefix}_done")])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

# Команда /start
@dp.message(Command("start"))
async def cmd_start(message: Message):
    if message.from_user.id not in ALLOWED_USERS and message.from_user.id != ADMIN_ID:
        await message.answer("⛔ У вас нет доступа к этому боту.")
        return
    
    await message.answer(
        "📊 *Система расчета роялти*\n\n"
        "Я помогу сформировать отчет по роялти.\n\n"
        "📌 *Доступные команды:*\n"
        "/report - сформировать отчет с выбором фильтров\n"
        "/help - справка\n\n"
        "Выберите действие:",
        parse_mode="Markdown"
    )

# Команда /help
@dp.message(Command("help"))
async def cmd_help(message: Message):
    await message.answer(
        "📖 *Справка*\n\n"
        "1. Нажмите /report\n"
        "2. Выберите договоры\n"
        "3. Выберите кварталы\n"
        "4. Выберите годы\n"
        "5. Выберите типы прав\n"
        "6. Укажите проценты для выбранных типов\n"
        "7. Выберите песни (можно пропустить)\n"
        "8. Получите отчет в виде текста и Excel-файла",
        parse_mode="Markdown"
    )

# Команда /report
@dp.message(Command("report"))
async def cmd_report(message: Message, state: FSMContext):
    if message.from_user.id not in ALLOWED_USERS and message.from_user.id != ADMIN_ID:
        await message.answer("⛔ У вас нет доступа к этому боту.")
        return
    
    contracts = get_unique_values("contract")
    if not contracts:
        await message.answer("❌ База данных пуста. Сначала загрузите данные.")
        return
    
    await state.update_data(selected_contracts=[], selected_quarters=[], selected_years=[], selected_types=[], selected_songs=[])
    await state.set_state(ReportState.waiting_contracts)
    
    keyboard = build_multi_select_keyboard(contracts, [], "contract")
    await message.answer("📋 *Выберите договоры:*\n(можно выбрать несколько)", parse_mode="Markdown", reply_markup=keyboard)

# Обработка callback'ов для выбора
@dp.callback_query()
async def handle_callback(callback: CallbackQuery, state: FSMContext):
    data = callback.data
    user_data = await state.get_data()
    
    # Обработка выбора договоров
    if data.startswith("contract_"):
        if data == "contract_done":
            selected = user_data.get("selected_contracts", [])
            if not selected:
                await callback.answer("Выберите хотя бы один договор!", show_alert=True)
                return
            
            # Переходим к выбору кварталов
            quarters = ["I", "II", "III", "IV"]
            await state.update_data(selected_quarters=[])
            await state.set_state(ReportState.waiting_quarters)
            keyboard = build_multi_select_keyboard(quarters, [], "quarter")
            await callback.message.edit_text("📅 *Выберите кварталы:*", parse_mode="Markdown", reply_markup=keyboard)
        
        elif data.startswith("contract_toggle_"):
            item = data.replace("contract_toggle_", "").replace("_", " ")
            selected = user_data.get("selected_contracts", [])
            if item in selected:
                selected.remove(item)
            else:
                selected.append(item)
            await state.update_data(selected_contracts=selected)
            
            contracts = get_unique_values("contract")
            page = user_data.get("contract_page", 0)
            keyboard = build_multi_select_keyboard(contracts, selected, "contract", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
        
        elif data.startswith("contract_page_"):
            page = int(data.split("_")[-1])
            await state.update_data(contract_page=page)
            contracts = get_unique_values("contract")
            selected = user_data.get("selected_contracts", [])
            keyboard = build_multi_select_keyboard(contracts, selected, "contract", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
    
    # Обработка выбора кварталов (аналогично)
    elif data.startswith("quarter_"):
        if data == "quarter_done":
            selected = user_data.get("selected_quarters", [])
            if not selected:
                await callback.answer("Выберите хотя бы один квартал!", show_alert=True)
                return
            
            # Переходим к выбору годов
            years = get_unique_values("year")
            await state.update_data(selected_years=[])
            await state.set_state(ReportState.waiting_years)
            keyboard = build_multi_select_keyboard(years, [], "year")
            await callback.message.edit_text("📆 *Выберите годы:*", parse_mode="Markdown", reply_markup=keyboard)
        
        elif data.startswith("quarter_toggle_"):
            item = data.replace("quarter_toggle_", "").replace("_", " ")
            selected = user_data.get("selected_quarters", [])
            if item in selected:
                selected.remove(item)
            else:
                selected.append(item)
            await state.update_data(selected_quarters=selected)
            
            quarters = ["I", "II", "III", "IV"]
            page = user_data.get("quarter_page", 0)
            keyboard = build_multi_select_keyboard(quarters, selected, "quarter", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
        
        elif data.startswith("quarter_page_"):
            page = int(data.split("_")[-1])
            await state.update_data(quarter_page=page)
            quarters = ["I", "II", "III", "IV"]
            selected = user_data.get("selected_quarters", [])
            keyboard = build_multi_select_keyboard(quarters, selected, "quarter", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
    
    # Обработка выбора годов
    elif data.startswith("year_"):
        if data == "year_done":
            selected = user_data.get("selected_years", [])
            if not selected:
                await callback.answer("Выберите хотя бы один год!", show_alert=True)
                return
            
            # Переходим к выбору типов прав
            types = ["Авторские", "Смежные"]
            await state.update_data(selected_types=[])
            await state.set_state(ReportState.waiting_types)
            keyboard = build_multi_select_keyboard(types, [], "type")
            await callback.message.edit_text("⚖️ *Выберите типы прав:*", parse_mode="Markdown", reply_markup=keyboard)
        
        elif data.startswith("year_toggle_"):
            item = data.replace("year_toggle_", "").replace("_", " ")
            selected = user_data.get("selected_years", [])
            if item in selected:
                selected.remove(item)
            else:
                selected.append(item)
            await state.update_data(selected_years=selected)
            
            years = get_unique_values("year")
            page = user_data.get("year_page", 0)
            keyboard = build_multi_select_keyboard(years, selected, "year", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
        
        elif data.startswith("year_page_"):
            page = int(data.split("_")[-1])
            await state.update_data(year_page=page)
            years = get_unique_values("year")
            selected = user_data.get("selected_years", [])
            keyboard = build_multi_select_keyboard(years, selected, "year", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
    
    # Обработка выбора типов прав
    elif data.startswith("type_"):
        if data == "type_done":
            selected = user_data.get("selected_types", [])
            if not selected:
                await callback.answer("Выберите хотя бы один тип!", show_alert=True)
                return
            
            # Если выбраны оба типа, запрашиваем проценты
            if "Авторские" in selected and "Смежные" in selected:
                await state.set_state(ReportState.waiting_author_percent)
                await callback.message.edit_text("💰 *Укажите процент блогера для АВТОРСКИХ прав:*\n(например: 50)", parse_mode="Markdown")
            elif "Авторские" in selected:
                await state.update_data(author_percent=0, related_percent=0)
                await state.set_state(ReportState.waiting_songs)
                songs = get_songs()
                keyboard = build_multi_select_keyboard(songs, [], "song")
                await callback.message.edit_text("🎵 *Выберите песни:*\n(можно выбрать несколько, или нажмите Готово чтобы взять все)", parse_mode="Markdown", reply_markup=keyboard)
            else:  # только смежные
                await state.update_data(author_percent=0, related_percent=0)
                await state.set_state(ReportState.waiting_related_percent)
                await callback.message.edit_text("💰 *Укажите процент блогера для СМЕЖНЫХ прав:*\n(например: 30)", parse_mode="Markdown")
        
        elif data.startswith("type_toggle_"):
            item = data.replace("type_toggle_", "").replace("_", " ")
            selected = user_data.get("selected_types", [])
            if item in selected:
                selected.remove(item)
            else:
                selected.append(item)
            await state.update_data(selected_types=selected)
            
            types = ["Авторские", "Смежные"]
            page = user_data.get("type_page", 0)
            keyboard = build_multi_select_keyboard(types, selected, "type", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
    
    # Обработка выбора песен
    elif data.startswith("song_"):
        if data == "song_done":
            await state.set_state(None)
            
            # Получаем все данные для отчета
            user_data = await state.get_data()
            await generate_report(callback.message, user_data)
        
        elif data.startswith("song_toggle_"):
            item = data.replace("song_toggle_", "").replace("_", " ")
            selected = user_data.get("selected_songs", [])
            if item in selected:
                selected.remove(item)
            else:
                selected.append(item)
            await state.update_data(selected_songs=selected)
            
            songs = get_songs()
            page = user_data.get("song_page", 0)
            keyboard = build_multi_select_keyboard(songs, selected, "song", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
        
        elif data.startswith("song_page_"):
            page = int(data.split("_")[-1])
            await state.update_data(song_page=page)
            songs = get_songs()
            selected = user_data.get("selected_songs", [])
            keyboard = build_multi_select_keyboard(songs, selected, "song", page)
            await callback.message.edit_reply_markup(reply_markup=keyboard)
    
    await callback.answer()

# Генерация отчета
async def generate_report(message, user_data):
    await message.answer("📊 Формирую отчет, пожалуйста подождите...")
    
    selected_contracts = user_data.get("selected_contracts", [])
    selected_quarters = user_data.get("selected_quarters", [])
    selected_years = user_data.get("selected_years", [])
    selected_types = user_data.get("selected_types", [])
    selected_songs = user_data.get("selected_songs", [])
    author_percent = user_data.get("author_percent", 0)
    related_percent = user_data.get("related_percent", 0)
    
    # Строим SQL-запрос
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM royalties WHERE 1=1"
    params = []
    
    if selected_contracts:
        query += f" AND contract IN ({','.join(['?']*len(selected_contracts))})"
        params.extend(selected_contracts)
    if selected_quarters:
        query += f" AND quarter IN ({','.join(['?']*len(selected_quarters))})"
        params.extend(selected_quarters)
    if selected_years:
        query += f" AND year IN ({','.join(['?']*len(selected_years))})"
        params.extend(selected_years)
    if selected_types:
        query += f" AND type IN ({','.join(['?']*len(selected_types))})"
        params.extend(selected_types)
    if selected_songs:
        query += f" AND display_name IN ({','.join(['?']*len(selected_songs))})"
        params.extend(selected_songs)
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        await message.answer("❌ Нет данных по выбранным фильтрам.")
        return
    
    # Расчет итогов
    TAX_RATE = 0.06  # 6%
    
    author_total_revenue = 0
    related_total_revenue = 0
    
    for row in rows:
        if row["type"] == "Авторские":
            author_total_revenue += row["sum"]
        else:
            related_total_revenue += row["sum"]
    
    # Расчет чистой выручки и выплат
    author_net = author_total_revenue * (1 - TAX_RATE)
    related_net = related_total_revenue * (1 - TAX_RATE)
    
    author_payout = author_net * (author_percent / 100) if author_percent > 0 else 0
    related_payout = related_net * (related_percent / 100) if related_percent > 0 else 0
    total_payout = author_payout + related_payout
    
    # Формируем текстовый отчет
    report_text = "📊 *ОТЧЕТ ПО РОЯЛТИ*\n\n"
    report_text += f"📋 *Договоры:* {', '.join(selected_contracts)}\n"
    report_text += f"📅 *Кварталы:* {', '.join(selected_quarters)}\n"
    report_text += f"📆 *Годы:* {', '.join(map(str, selected_years))}\n"
    report_text += f"⚖️ *Типы прав:* {', '.join(selected_types)}\n\n"
    
    if selected_types:
        if "Авторские" in selected_types:
            report_text += "💰 *АВТОРСКИЕ ПРАВА*"
            if author_percent > 0:
                report_text += f" (процент блогера: {author_percent}%)\n"
            else:
                report_text += "\n"
            report_text += f"  Общий доход: {author_total_revenue:,.2f} ₽\n"
            report_text += f"  Налог (6%): {author_total_revenue * TAX_RATE:,.2f} ₽\n"
            report_text += f"  Чистая выручка: {author_net:,.2f} ₽\n"
            if author_percent > 0:
                report_text += f"  К выплате блогеру: {author_payout:,.2f} ₽\n"
            report_text += "\n"
        
        if "Смежные" in selected_types:
            report_text += "🎵 *СМЕЖНЫЕ ПРАВА*"
            if related_percent > 0:
                report_text += f" (процент блогера: {related_percent}%)\n"
            else:
                report_text += "\n"
            report_text += f"  Общий доход: {related_total_revenue:,.2f} ₽\n"
            report_text += f"  Налог (6%): {related_total_revenue * TAX_RATE:,.2f} ₽\n"
            report_text += f"  Чистая выручка: {related_net:,.2f} ₽\n"
            if related_percent > 0:
                report_text += f"  К выплате блогеру: {related_payout:,.2f} ₽\n"
            report_text += "\n"
    
    report_text += f"📌 *ИТОГО К ВЫПЛАТЕ:* {total_payout:,.2f} ₽\n\n"
    report_text += "📎 *Детализация в прикрепленном Excel-файле*"
    
    # Генерация Excel
    excel_filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    
    # Лист с детализацией
    ws = wb.active
    ws.title = "Детализация"
    
    headers = ["Тип", "Договор", "Квартал", "Год", "Название", "Авторы/Исполнители", "Сумма", "Налог (6%)", "Чистая выручка", "Процент блогера", "К выплате"]
    ws.append(headers)
    
    for row in rows:
        tax = row["sum"] * TAX_RATE
        net = row["sum"] - tax
        if row["type"] == "Авторские":
            percent = author_percent
            payout = net * (percent / 100) if percent > 0 else 0
        else:
            percent = related_percent
            payout = net * (percent / 100) if percent > 0 else 0
        
        additional = row["additional_info"] if row["additional_info"] else ""
        ws.append([
            row["type"], row["contract"], row["quarter"], row["year"],
            row["display_name"], additional, row["sum"], tax, net, percent, payout
        ])
    
    # Лист со сводкой
    ws_summary = wb.create_sheet("Сводка")
    ws_summary.append(["Показатель", "Значение"])
    ws_summary.append(["Договоры", ", ".join(selected_contracts)])
    ws_summary.append(["Кварталы", ", ".join(selected_quarters)])
    ws_summary.append(["Годы", ", ".join(map(str, selected_years))])
    ws_summary.append(["Типы прав", ", ".join(selected_types)])
    ws_summary.append([])
    
    if "Авторские" in selected_types:
        ws_summary.append(["АВТОРСКИЕ ПРАВА", ""])
        ws_summary.append(["Общий доход", f"{author_total_revenue:,.2f} ₽"])
        ws_summary.append(["Налог (6%)", f"{author_total_revenue * TAX_RATE:,.2f} ₽"])
        ws_summary.append(["Чистая выручка", f"{author_net:,.2f} ₽"])
        ws_summary.append(["Процент блогера", f"{author_percent}%"])
        ws_summary.append(["К выплате", f"{author_payout:,.2f} ₽"])
        ws_summary.append([])
    
    if "Смежные" in selected_types:
        ws_summary.append(["СМЕЖНЫЕ ПРАВА", ""])
        ws_summary.append(["Общий доход", f"{related_total_revenue:,.2f} ₽"])
        ws_summary.append(["Налог (6%)", f"{related_total_revenue * TAX_RATE:,.2f} ₽"])
        ws_summary.append(["Чистая выручка", f"{related_net:,.2f} ₽"])
        ws_summary.append(["Процент блогера", f"{related_percent}%"])
        ws_summary.append(["К выплате", f"{related_payout:,.2f} ₽"])
        ws_summary.append([])
    
    ws_summary.append(["ИТОГО К ВЫПЛАТЕ", f"{total_payout:,.2f} ₽"])
    
    wb.save(excel_filename)
    
    # Отправляем текстовый отчет и файл
    await message.answer(report_text, parse_mode="Markdown")
    await message.answer_document(FSInputFile(excel_filename), caption="📎 Детализация отчета")
    
    # Удаляем временный файл
    os.remove(excel_filename)

# Обработка текстовых сообщений (для ввода процентов)
@dp.message(ReportState.waiting_author_percent)
async def process_author_percent(message: Message, state: FSMContext):
    try:
        percent = float(message.text.replace(",", "."))
        if percent < 0 or percent > 100:
            await message.answer("❌ Процент должен быть от 0 до 100. Попробуйте еще раз.")
            return
        await state.update_data(author_percent=percent)
        
        selected_types = (await state.get_data()).get("selected_types", [])
        if "Смежные" in selected_types:
            await state.set_state(ReportState.waiting_related_percent)
            await message.answer("💰 *Укажите процент блогера для СМЕЖНЫХ прав:*\n(например: 30)", parse_mode="Markdown")
        else:
            await state.update_data(related_percent=0)
            await state.set_state(ReportState.waiting_songs)
            songs = get_songs()
            keyboard = build_multi_select_keyboard(songs, [], "song")
            await message.answer("🎵 *Выберите песни:*\n(можно выбрать несколько, или нажмите Готово чтобы взять все)", parse_mode="Markdown", reply_markup=keyboard)
    except ValueError:
        await message.answer("❌ Пожалуйста, введите число (например: 50 или 33.5)")

@dp.message(ReportState.waiting_related_percent)
async def process_related_percent(message: Message, state: FSMContext):
    try:
        percent = float(message.text.replace(",", "."))
        if percent < 0 or percent > 100:
            await message.answer("❌ Процент должен быть от 0 до 100. Попробуйте еще раз.")
            return
        await state.update_data(related_percent=percent)
        await state.set_state(ReportState.waiting_songs)
        songs = get_songs()
        keyboard = build_multi_select_keyboard(songs, [], "song")
        await message.answer("🎵 *Выберите песни:*\n(можно выбрать несколько, или нажмите Готово чтобы взять все)", parse_mode="Markdown", reply_markup=keyboard)
    except ValueError:
        await message.answer("❌ Пожалуйста, введите число (например: 30 или 25.5)")

# Запуск бота и Flask
def run_flask():
    port = int(os.environ.get('PORT', 10000))
    flask_app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    # Запускаем Flask в отдельном потоке
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()
    
    # Запускаем бота
    asyncio.run(main())
